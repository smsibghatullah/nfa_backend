import openpyxl, json
import datetime
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from .models import (Candidate, JobPost, TestSchedule, 
                     ContactRequest, Document, 
                     Profile, JobListing, JobApplication, JobQuestion, ApplicationDocument)
from .serializers import (ContactRequestSerializer, DocumentSerializer, ProfileSerializer, 
                          JobListingSerializer, JobApplicationSerializer, JobApplicationReviewSerializer, UploadApplicationDocumentSerializer)
from .utils import calculate_age, get_highest_qualification, calculate_total_experience, QUALIFICATION_ORDER

from rest_framework.decorators import api_view, parser_classes, permission_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.permissions import AllowAny, IsAuthenticated

@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
@permission_classes([AllowAny])
def contact_us(request):
    serializer = ContactRequestSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Contact request submitted successfully."},
            status=status.HTTP_201_CREATED
        )
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([AllowAny])
def get_documents(request):
    documents = Document.objects.all()
    serializer = DocumentSerializer(documents, many=True, context={'request': request})
    return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
@permission_classes([AllowAny])
def upload_document(request):
    serializer = DocumentSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


def parse_excel_date(value):
    if isinstance(value, datetime.date):
        return value
    value_str = str(value).strip()
    for fmt in ("%d %b %Y", "%d %B %Y", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(value_str, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Invalid date format: {value_str}")


@staff_member_required
def upload_schedule(request):
    if request.method == 'POST' and request.FILES.get('xlsx_file'):
        file = request.FILES['xlsx_file']

        if not file.name.endswith('.xlsx'):
            messages.error(request, "Invalid file format. Please upload a .xlsx file.")
            return redirect(request.path)

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception as e:
            messages.error(request, f"Error reading Excel file: {str(e)}")
            return redirect(request.path)

        expected_columns = [
            'Sr.No.', 'Roll No', 'Name', 'Father Name', 'CNIC', 'Post Applied For',
            'Postal Address', 'Mobile No.', 'Paper', 'Test Date', 'Session',
            'Reporting Time', 'Conduct Time', 'Venue'
        ]

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if headers != expected_columns:
            messages.error(
                request,
                "Excel format is invalid. Ensure headers match the required structure."
            )
            return redirect(request.path)

        for row in ws.iter_rows(min_row=2, values_only=True):
            if all((cell is None or (isinstance(cell, str) and cell.strip() == "")) for cell in row):
                continue
            try:
                sr_no, roll_no, name, father_name, cnic, post_title, postal_address, \
                mobile_no, paper, test_date, session, reporting_time, conduct_time, venue = row

                test_date = parse_excel_date(test_date)

                code_value = post_title.strip().upper().replace(" ", "_")[:20]

                job_post, _ = JobPost.objects.update_or_create(
                    code=code_value,
                    defaults={"title": post_title}
                )

                candidate, _ = Candidate.objects.update_or_create(
                    roll_no=roll_no,
                    defaults={
                        'name': name,
                        'father_name': father_name,
                        'cnic': cnic,
                        'postal_address': postal_address,
                        'mobile_no': mobile_no
                    }
                )

                TestSchedule.objects.create(
                    candidate=candidate,
                    job_post=job_post,
                    paper=paper,
                    test_date=test_date,
                    session=session,
                    reporting_time=reporting_time,
                    conduct_time=conduct_time,
                    venue=venue
                )

            except Exception as e:
                messages.error(request, f"Error processing row: {row}. Error: {str(e)}")

        messages.success(request, "File uploaded successfully.")
        return redirect(request.path)

    return render(request, 'admin/candidates/upload.html')

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_my_profile(request):
    try:
        profile = request.user.profile
    except Profile.DoesNotExist:
        return Response({"detail": "Profile not found."}, status=status.HTTP_404_NOT_FOUND)
    serializer = ProfileSerializer(profile, context={'request': request})
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_profile(request):
    data = request.data.copy()
    data['user'] = request.user.id
    serializer = ProfileSerializer(data=data, context={'request': request})
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT', 'PATCH'])
@permission_classes([IsAuthenticated])
def update_profile(request, pk):
    try:
        profile = Profile.objects.get(pk=pk, user=request.user)
    except Profile.DoesNotExist:
        return Response({"detail": "Profile not found."}, status=status.HTTP_404_NOT_FOUND)
    serializer = ProfileSerializer(profile, data=request.data, partial=True, context={'request': request})
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([AllowAny])
def list_job_listings(request):
    listings = JobListing.objects.all()
    serializer = JobListingSerializer(listings, many=True, context={'request': request})
    return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([AllowAny])
def retrieve_job_listing(request, pk):
    try:
        listing = JobListing.objects.get(pk=pk)
    except JobListing.DoesNotExist:
        return Response({"detail": "Job listing not found."}, status=status.HTTP_404_NOT_FOUND)

    serializer = JobListingSerializer(listing, context={'request': request})
    return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def application_eligibility_check(request):
    job_id = request.data.get("job_id")
    try:
        job_listing = JobListing.objects.get(pk=job_id)
    except JobListing.DoesNotExist:
        return Response({"eligible": False, "reason": "Job not found"}, status=404)

    try:
        profile = request.user.profile
    except Profile.DoesNotExist:
        return Response({"eligible": False, "reason": "Profile not found"}, status=400)

    age = calculate_age(profile.date_of_birth)
    if job_listing.minimum_age and age < job_listing.minimum_age:
        return Response({
            "eligible": False,
            "reason": f"You are not eligible for this position. Reason: Minimum age required is {job_listing.minimum_age}, your age is {age}."
        })

    highest_degree = get_highest_qualification(profile)
    if job_listing.minimum_qualification:
        required_level = QUALIFICATION_ORDER.get(job_listing.minimum_qualification, 0)
        applicant_level = QUALIFICATION_ORDER.get(highest_degree, 0)
        if applicant_level < required_level:
            return Response({
                "eligible": False,
                "reason": f"You are not eligible for this position. Reason: Required qualification is {job_listing.minimum_qualification}, your highest qualification is {highest_degree or 'None'}."
            })

    total_exp_years = calculate_total_experience(profile)
    if job_listing.required_experience and total_exp_years < job_listing.required_experience:
        return Response({
            "eligible": False,
            "reason": f"You are not eligible for this position. Reason: Required experience is {job_listing.required_experience} years, your total is {total_exp_years} years."
        })

    return Response({
        "eligible": True,
        "reason": "Eligible. You may proceed with the application form."
    })

@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
@permission_classes([IsAuthenticated])
def upload_application_file(request):
    file_obj = request.FILES.get('file')
    name = request.data.get('name') or (file_obj.name if file_obj else None)

    if not file_obj:
        return Response({"error": "No file provided"}, status=400)

    document = ApplicationDocument.objects.create(
        name=name,
        file=file_obj 
    )

    serializer = UploadApplicationDocumentSerializer(document, context={'request': request})
    return Response(serializer.data, status=201)


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
@permission_classes([IsAuthenticated])
def create_job_application(request):
    try:
        profile = request.user.profile
    except Profile.DoesNotExist:
        return Response({"error": "User does not have a profile"}, status=400)

    data = request.data.copy()
    answers_data = json.loads(data.get('answers', '[]'))
    document_ids = json.loads(data.get('document_ids', '[]'))

    serializer = JobApplicationSerializer(
        data={
            'job_listing': data.get('job_listing'),
            'answers': answers_data,
            'document_ids': document_ids
        },
        context={'request': request}
    )

    if serializer.is_valid():
        application = serializer.save(applicant=profile)
        return Response(JobApplicationReviewSerializer(application, context={'request': request}).data, status=201)
    else:
        return Response(serializer.errors, status=400)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def review_job_application(request, application_id):
    try:
        application = JobApplication.objects.get(pk=application_id, applicant=request.user.profile)
    except JobApplication.DoesNotExist:
        return Response({"detail": "Application not found."}, status=404)

    serializer = JobApplicationReviewSerializer(application, context={'request': request})
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def confirm_job_application(request, application_id):
    try:
        application = JobApplication.objects.get(pk=application_id, applicant=request.user.profile)
    except JobApplication.DoesNotExist:
        return Response({"detail": "Application not found."}, status=404)

    if application.is_confirmed:
        return Response({"detail": "Application already confirmed."}, status=400)

    application.is_confirmed = True
    application.save()
    return Response({"message": "Application confirmed successfully."})